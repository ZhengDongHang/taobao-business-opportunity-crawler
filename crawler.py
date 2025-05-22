#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
second_gen_crawler.py：按照 Excel 指示，对准备爬取的一级类目执行爬虫
CSV 重复跳过；并行地对各一级分类分别启动浏览器执行爬取，并跳过已经爬取的类目
一级分类全部文件爬取完毕后各自修改 Excel 第三列
"""

import os
import csv
import time
import subprocess
from multiprocessing import Pool, cpu_count

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from config import category, is_head, processing_quantity

# ———— 配置 ————
PAGE_URL         = ("https://air.1688.com/app/fuwu-assets/work-buyer-plugin-frame"
                    "/home.html?spm=a260k.home2024.leftmenu_EXPEND."
                    "ddefault0xiayoushangji0of0shangpin.663335e4TMrI8K")
OUTPUT_XLSX      = "一级类目.xlsx"
CHROMEDRIVER_PATH = os.path.join(os.path.dirname(__file__),
                                  "chromedriver-win64", "chromedriver.exe")
# 并行进程数
PROCESS_COUNT    = min(processing_quantity, cpu_count())
# —————————————————

def find_chrome_path():
    try:
        result = subprocess.check_output("where chrome", shell=True, encoding="utf-8")
        return result.strip().splitlines()[0]
    except Exception:
        return r"C:\Program Files\Google\Chrome\Application\chrome.exe"

CHROME_EXE = find_chrome_path()
print(f"[信息] Chrome 路径自动识别为: {CHROME_EXE}")

def open_category_popup(driver):
    try:
        sel = driver.find_element(By.CSS_SELECTOR, ".next-channel-select-inner")
        if sel.get_attribute("aria-expanded") != "true":
            ActionChains(driver).move_to_element(sel).click().perform()
            time.sleep(1)
    except:
        pass

def scrape_leaf(driver, folder, label):
    os.makedirs(folder, exist_ok=True)
    out = os.path.join(folder, f"{label.replace('/', '_')}.csv")
    # CSV 已存在，就跳过
    if os.path.exists(out):
        return

    print(f"  -> [叶子] 抓取分类 {label}")
    results = []
    try:
        pg = driver.find_element(By.CSS_SELECTOR, "span.next-channel-pagination-display")
        total = int(pg.text.split("/")[1])
    except:
        total = 1

    for p in range(1, total + 1):
        print(f"    第 {p}/{total} 页", end="")
        imgs = driver.find_elements(
            By.CSS_SELECTOR,
            "tbody.next-channel-table-body tr.next-channel-table-row img.offer-img"
        )
        for img in imgs:
            try:
                href = img.find_element(By.XPATH, "./ancestor::a").get_attribute("href")
                src  = img.get_attribute("src")
                results.append((href, src))
            except:
                pass
        if p < total:
            try:
                btn = driver.find_element(By.XPATH, f"//button[span[text()='{p+1}']]")
                btn.click()
                time.sleep(0.8)
            except:
                pass
        print(f" 累计 {len(results)} 条")

    # 写CSV
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Link", "Image"])
        w.writerows(results)
    print(f"    ✅ 已保存 {out}")

def scrape_products(driver, second_text, third_items, first_level_folder):
    base_folder = os.path.join("product_link", first_level_folder,
                               second_text.replace("/", "_"))
    os.makedirs(base_folder, exist_ok=True)

    texts = [li.find_element(By.CSS_SELECTOR, ".next-channel-menu-item-text").text.strip()
             for li in third_items]
    for third_text in texts:
        print(f"  -> 抓取 {second_text} > {third_text}")
        open_category_popup(driver)
        wrappers = driver.find_elements(
            By.CSS_SELECTOR,
            ".next-channel-overlay-wrapper .next-channel-cascader-menu-wrapper"
        )
        # 选二级
        for li in wrappers[1].find_elements(By.CSS_SELECTOR, "li.next-channel-menu-item"):
            if li.find_element(By.CSS_SELECTOR, ".next-channel-menu-item-text").text.strip() == second_text:
                try:
                    WebDriverWait(driver,5).until(EC.element_to_be_clickable(li))
                    li.click(); time.sleep(0.5)
                except:
                    pass
                break
        # 选三级
        wrappers = driver.find_elements(
            By.CSS_SELECTOR,
            ".next-channel-overlay-wrapper .next-channel-cascader-menu-wrapper"
        )
        target = None
        for li in wrappers[2].find_elements(By.CSS_SELECTOR, "li.next-channel-menu-item"):
            if li.find_element(By.CSS_SELECTOR, ".next-channel-menu-item-text").text.strip() == third_text:
                target = li; break
        if not target:
            print(f"    [跳过] 找不到三级分类 {third_text}")
            continue
        try:
            WebDriverWait(driver,5).until(EC.element_to_be_clickable(target))
            target.click(); time.sleep(0.5)
        except Exception as e:
            print(f"    [跳过] 点击三级分类失败：{e}")
            continue

        scrape_leaf(driver, base_folder, third_text)

def select_categories(driver, first_level_folder):
    open_category_popup(driver)
    wrappers = driver.find_elements(
        By.CSS_SELECTOR,
        ".next-channel-overlay-wrapper .next-channel-cascader-menu-wrapper"
    )
    if len(wrappers) < 2:
        print("Error: 找不到二级菜单容器")
        return

    second_items = wrappers[1].find_elements(By.CSS_SELECTOR, "li.next-channel-menu-item")
    for li in second_items:
        second_text = li.find_element(By.CSS_SELECTOR, ".next-channel-menu-item-text").text.strip()
        print(f"\n=== 处理二级分类：{second_text} ===")
        try:
            WebDriverWait(driver,5).until(EC.element_to_be_clickable(li))
            li.click(); time.sleep(0.5)
        except Exception as e:
            print(f"  [跳过] 点击二级分类失败：{e}")
            open_category_popup(driver)
            continue

        wrappers = driver.find_elements(
            By.CSS_SELECTOR,
            ".next-channel-overlay-wrapper .next-channel-cascader-menu-wrapper"
        )
        # 无第三级
        if len(wrappers) < 3 or not wrappers[2].find_elements(By.CSS_SELECTOR, "li.next-channel-menu-item"):
            folder = os.path.join("product_link", first_level_folder,
                                  second_text.replace("/", "_"))
            scrape_leaf(driver, folder, second_text)
        else:
            third_items = wrappers[2].find_elements(By.CSS_SELECTOR, "li.next-channel-menu-item")
            scrape_products(driver, second_text, third_items, first_level_folder)

        open_category_popup(driver)

def worker(args):
    """每个进程处理一个一级类目"""
    cat_name, row_idx = args
    driver_opts = Options()
    if is_head == True:
        pass
    else:
        driver_opts.add_argument("--headless")  # ✅无头
    driver_opts.add_argument("--no-sandbox")
    driver_opts.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(service=Service(CHROMEDRIVER_PATH), options=driver_opts)

    try:
        # 打开首页
        driver.get(PAGE_URL); time.sleep(2)

        # 切换榜单
        try:
            rg = driver.find_element(By.CSS_SELECTOR, "div.next-channel-radio-group")
            for lbl in rg.find_elements(By.CSS_SELECTOR, "span.next-channel-radio-label"):
                if lbl.text.strip() == "趋势热品":
                    lbl.click(); time.sleep(0.5); break
        except:
            pass

        # 切 tabs
        WebDriverWait(driver,10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "ul.next-channel-tabs-nav"))
        )
        tab = driver.find_element(
            By.XPATH,
            f"//ul[contains(@class,'next-channel-tabs-nav')]"
            f"/li[.//div[text()='{category}']]"
        )
        tab.click(); time.sleep(0.5)

        print(f"\n>> 开始一级类目：{cat_name}")
        first_folder = cat_name.replace("/", "_")
        # 点击一级分类
        open_category_popup(driver)
        wrappers = driver.find_elements(
            By.CSS_SELECTOR,
            ".next-channel-overlay-wrapper .next-channel-cascader-menu-wrapper"
        )
        for li in wrappers[0].find_elements(By.CSS_SELECTOR, "li.next-channel-menu-item"):
            if li.get_attribute("title").strip() == cat_name:
                WebDriverWait(driver,5).until(EC.element_to_be_clickable(li))
                li.click(); time.sleep(0.5)
                break

        # 爬二三级
        select_categories(driver, first_folder)

        # 回写 Excel “完成” 标记
        wb = load_workbook(OUTPUT_XLSX)
        ws = wb.active
        ws.cell(row=row_idx, column=3, value=1)
        wb.save(OUTPUT_XLSX)
        print(f"✅ 完成一级类目：{cat_name}")

    finally:
        driver.quit()

def main():
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active

    tasks = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        name = row[0].value
        ready = row[1].value
        if ready == 1:
            tasks.append((name, i))

    if not tasks:
        print("没有待爬取的一级类目。")
        return

    with Pool(PROCESS_COUNT) as pool:
        pool.map(worker, tasks)
if __name__ == '__main__':
    main()

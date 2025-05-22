#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
final_process.py：对解析之后的数据计算 ‘总访客数’，‘手淘占比’，删除重复数据行，并冻结表头
           —跳过读取失败的损坏文件并记录—
"""


import sys
import os
import csv
import zipfile
from openpyxl import load_workbook

# 配置项
INPUT_DIR = "output"
BAD_RECORD_CSV = "损坏文件记录.csv"

def safe_float(value):
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except:
        return None

def safe_float(val):
    try:
        return float(str(val).replace('%', '').strip())
    except (ValueError, TypeError):
        return None

def process_excel(path):
    wb = load_workbook(path)
    ws = wb.active

    # 冻结第一行
    ws.freeze_panes = "A2"

    # 确保“总访客数”“手淘占比”列存在
    headers = [cell.value for cell in ws[1]]
    if "总访客数" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="总访客数")
        headers.append("总访客数")
    if "手淘占比" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="手淘占比")
        headers.append("手淘占比")

    # 列名到列号映射
    col_map = {name: idx + 1 for idx, name in enumerate(headers)}
    idx_clicks         = col_map["点击人气"]
    idx_index          = col_map["商品指数"]
    idx_rate           = col_map["支付转化率"]
    idx_total_visitors = col_map["总访客数"]
    idx_taobao_ratio   = col_map["手淘占比"]
    idx_search_hot = col_map.get("搜索热度")

    seen = set()
    to_delete = []

    # 从下往上遍历以便删除
    for row in reversed(list(ws.iter_rows(min_row=2, max_row=ws.max_row))):
        r = row[0].row
        # 唯一签名：除图片外所有单元格值
        sig = tuple(
            (cell.value or "").strip() if isinstance(cell.value, str) else cell.value
            for cell in row[1:]
        )
        if sig in seen:
            to_delete.append(r)
            continue
        seen.add(sig)

        # 重新计算“总访客数” & “手淘占比”
        clicks = safe_float(ws.cell(r, idx_clicks).value)
        index  = safe_float(ws.cell(r, idx_index).value)
        rate   = safe_float(ws.cell(r, idx_rate).value)

        total_visitors = index / (rate / 100) if rate and rate != 0 else None
        taobao_ratio   = clicks / total_visitors if total_visitors else None

        ws.cell(row=r, column=idx_total_visitors,
                value=round(total_visitors, 2) if total_visitors else None)
        ws.cell(row=r, column=idx_taobao_ratio,
                value=round(100 * taobao_ratio, 4) if taobao_ratio else None)
        # 去除“搜索热度”列中的百分号
        if idx_search_hot:
            val_h = ws.cell(row=r, column=idx_search_hot).value
            if isinstance(val_h, str) and val_h.endswith("%"):
                val_h = val_h.replace("%", "").strip()
                ws.cell(row=r, column=idx_search_hot, value=val_h)

    # 给“手淘占比”和“飙升热度（%）”所有非空值统一添加百分号
    idx_soaring = col_map.get("飙升热度（%）")
    for r in range(2, ws.max_row + 1):
        # 手淘占比
        val_t = ws.cell(row=r, column=idx_taobao_ratio).value
        if val_t is not None:
            s = str(val_t)
            if not s.endswith("%"):
                ws.cell(row=r, column=idx_taobao_ratio, value=f"{s}%")
        # 飙升热度
        if idx_soaring:
            val_s = ws.cell(row=r, column=idx_soaring).value
            if val_s is not None:
                s2 = str(val_s)
                if not s2.endswith("%"):
                    ws.cell(row=r, column=idx_soaring, value=f"{s2}%")

    wb.save(path)
    print(f"已处理：{path}")

def main():
    bad_records = []  # list of (二级目录, 文件名)

    for root, _, files in os.walk(INPUT_DIR):
        for fname in files:
            if not fname.lower().endswith(".xlsx") or fname.startswith("~$"):
                continue
            full_path = os.path.join(root, fname)
            try:
                process_excel(full_path)
            except (zipfile.BadZipFile, KeyError, Exception) as e:
                # 记录并跳过
                rel_dir = os.path.relpath(root, INPUT_DIR)
                bad_records.append((rel_dir, fname))
                print(f"[损坏跳过] {os.path.join(rel_dir, fname)}：{e}")
                continue

    # 写出坏文件记录
    if bad_records:
        with open(BAD_RECORD_CSV, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["二级目录", "文件名称"])
            writer.writerows(bad_records)
        print(f"已生成坏文件记录：{BAD_RECORD_CSV}")

if __name__ == '__main__':
    main()
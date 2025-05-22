#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pipeline.py：解析数据，读取 CSV，多进程可视模式爬取指标，下载原图并写入 Excel
         （并行处理每个 CSV + 顺序爬指标+下载原图+跳过已生成文件+图片缩放显示）
"""

from PIL import JpegImagePlugin
JpegImagePlugin._getmp = lambda  x:None

import os
import sys
import time
import random
import requests
import pandas as pd
from io import BytesIO
from multiprocessing import Pool

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from config import candidate_headers,time_range, processing_quantity, is_head
from requests.exceptions import ReadTimeout
from selenium.common.exceptions import WebDriverException


# -------- 配置 --------
INPUT_DIR   = "product_link"
OUTPUT_DIR  = "output"
PROCESSES   = processing_quantity   # 并行处理 CSV 的进程数

# 插入到 Excel 中的尺寸（像素）
IMAGE_SHOW_WIDTH  = 50
IMAGE_SHOW_HEIGHT = 50

# Excel 单元格设置
IMAGE_COL = "A"
COL_WIDTH = 10   # 大概对应 10 个字符宽度 ≈ 75px
ROW_HEIGHT = 40  # 大概 40 磅 ≈ 53px
# ----------------------


def init_driver():
    """无头 Chrome Driver，用于多进程调用"""
    base = os.path.dirname(os.path.abspath(__file__))
    driver_path = os.path.join(base, "chromedriver-win64", "chromedriver.exe")
    opts = Options()
    if is_head == True:
        pass
    else:
        opts.add_argument("--headless")  # ✅无头
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    service = Service(executable_path=driver_path)
    return webdriver.Chrome(service=service, options=opts)


def crawl_metrics(driver, url, time_range):
    """单次：加载页面、切换30天、解析指标/商品名/详情链接"""
    # 1) 加载页面，最多6次重试
    for attempt in range(1, 7):
        try:
            driver.get(url)
            break
        except WebDriverException as e:
            print(f"[警告] 第{attempt}次加载页面失败: {e}")
            time.sleep(5)
    else:
        print(f"[错误] 无法加载页面，跳过：{url}")
        keys = [
            "搜索人气","搜索热度","点击人气","点击热度",
            "点击率","交易指数","支付转化率","商品指数","飙升热度（%）",
            "商品名称","链接"
        ]
        return {k: None for k in keys}

    # 2) 切换近30天（如果需要），最多6次
    if time_range == '30':
        for attempt in range(1, 7):
            try:
                btn = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//button[.//span[contains(text(),'近30天')]]")
                    )
                )
                btn.click()
                WebDriverWait(driver, 5).until_not(
                    EC.text_to_be_present_in_element(
                        (By.CLASS_NAME, "card_content_item_box_number"), "加载中"
                    )
                )
                break
            except Exception as e:
                print(f"[警告] 第{attempt}次切换30天失败: {e}")
                time.sleep(1)
        time.sleep(1)  # 确保刷新到30天的数据

    # 3) 解析页面
    soup = BeautifulSoup(driver.page_source, "lxml")

    # 指标字段
    metric_keys = [
        "搜索人气","搜索热度","点击人气","点击热度",
        "点击率","交易指数","支付转化率","商品指数","飙升热度（%）"
    ]
    metrics = {k: None for k in metric_keys}

    for box in soup.select(".card_content_item_box"):
        t = box.select_one(".card_content_item_box_title")
        v = box.select_one(".card_content_item_box_number")
        if not t or not v:
            continue
        title = t.get_text(strip=True)
        val   = v.get_text(strip=True)
        for k in metric_keys:
            if title.startswith(k):
                metrics[k] = val
                break

    # 商品名称
    name_tag = soup.select_one("div.header_right_title")
    metrics["商品名称"] = name_tag.get_text(strip=True) if name_tag else None

    # 查看淘宝详情链接
    a = soup.find('a', string=lambda s: s and '查看淘宝商品详情' in s)
    metrics["链接"] = a['href'] if a and a.has_attr('href') else None

    return metrics


def crawl_metrics_with_retry(driver, url, time_range, retries=3):
    """重试调用 crawl_metrics，直到获得任一有效字段"""
    for i in range(1, retries + 1):
        data = crawl_metrics(driver, url, time_range)
        if any(v is not None for v in data.values()):
            return data
        print(f"[重试{i}] URL={url}")
    print(f"[失败] 放弃 URL={url}")
    return data


def convert_webp_to_png(stream: BytesIO) -> BytesIO:
    img = PILImage.open(stream).convert("RGB")
    out = BytesIO()
    img.save(out, format="PNG")
    out.seek(0)
    return out


def download_image_stream(image_url: str) -> BytesIO:
    """
    下载原图（或将 .webp / .mpo 转为 PNG 或 JPEG），不做额外压缩。
    返回 BytesIO 或 None。
    """
    if not image_url:
        return None

    for _ in range(3):  # 最多重试3次
        try:
            headers = random.choice(candidate_headers)
            r = requests.get(image_url, headers=headers, timeout=10)
            r.raise_for_status()
            ct = r.headers.get("Content-Type", "").lower()
            stream = BytesIO(r.content)

            # 处理 webp 图像
            if ".webp" in image_url.lower() or "webp" in ct:
                try:
                    stream = convert_webp_to_png(stream)
                    return stream
                except Exception as e:
                    print(f"[跳过] WEBP 图片转换失败：{image_url}，错误：{e}")
                    return None

            # 处理 mpo 图像
            if ".mpo" in image_url.lower() or "mpo" in ct:
                try:
                    img = PILImage.open(stream).convert("RGB")
                    out = BytesIO()
                    img.save(out, format="JPEG")
                    out.seek(0)
                    return out
                except Exception as e:
                    print(f"[跳过] MPO 图片解析失败：{image_url}，错误：{e}")
                    return None

            return stream

        except ReadTimeout:
            print(f"[重试] 请求超时：{image_url}")
            continue
        except Exception as e:
            print(f"[跳过] 图片下载失败：{image_url}，错误：{e}")
            return None

    print(f"[失败] 多次重试后仍无法下载：{image_url}")
    return None


def process_csv(input_csv: str, output_xlsx: str):
    """顺序处理单个 CSV，并写入 Excel"""
    driver = init_driver()
    try:
        df = pd.read_csv(input_csv)
        wb = Workbook()
        ws = wb.active

        # 新增 “图片链接” 列
        headers = [
            "图片", "链接", "图片链接", "商品名称",
            "搜索人气", "搜索热度", "点击人气", "点击热度",
            "点击率", "交易指数", "支付转化率", "商品指数", "飙升热度（%）"
        ]
        ws.append(headers)
        # 设置图片列宽和默认行高
        ws.column_dimensions[IMAGE_COL].width = COL_WIDTH

        row_no = 2
        for idx, row in df.iterrows():
            page_url  = row.get("Link", "").strip()
            image_url = row.get("Image", "").strip()
            if not page_url:
                continue

            data = crawl_metrics_with_retry(driver, page_url, time_range)
            # 检查除前三列外是否都有值
            if any(data.get(col) is None for col in headers[3:]):
                print(f"[跳过] 第{idx}行数据不完整，URL={page_url}")
                continue

            # 下载图片流
            img_stream = download_image_stream(image_url)
            if img_stream:
                try:
                    img = XLImage(img_stream)
                    # 设置在 Excel 中显示大小（像素）
                    img.width  = IMAGE_SHOW_WIDTH
                    img.height = IMAGE_SHOW_HEIGHT
                    # 调整对应行高
                    ws.row_dimensions[row_no].height = ROW_HEIGHT
                    ws.add_image(img, f"{IMAGE_COL}{row_no}")
                except Exception as e:
                    print(f"[警告] 插入图片失败 idx={idx}: {e}")

            # 写入 “链接”、“图片链接”、“商品名称”
            ws.cell(row=row_no, column=2, value=data.get("链接"))
            ws.cell(row=row_no, column=3, value=image_url)
            ws.cell(row=row_no, column=4, value=data.get("商品名称"))

            # 写入其它字段
            for i, key in enumerate(headers[4:], start=5):
                val = data.get(key)
                # 对“搜索热度”和“手淘占比”后面加"%"
                if key in ("搜索热度", "飙升热度（%）") and val is not None:
                    val = f"{val}%"
                ws.cell(row=row_no, column=i, value=val)

            row_no += 1

        os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
        wb.save(output_xlsx)
        print(f"✅ 已生成：{output_xlsx}")
    finally:
        driver.quit()


def worker(task):
    """子进程入口：task = (in_csv, out_xlsx)"""
    in_csv, out_xlsx = task
    if os.path.exists(out_xlsx):
        print(f"[跳过] 已存在：{out_xlsx}")
    else:
        process_csv(in_csv, out_xlsx)


def main():
    tasks = []
    for root, _, files in os.walk(INPUT_DIR):
        for fname in files:
            if not fname.lower().endswith(".csv"):
                continue
            in_csv = os.path.join(root, fname)
            rel    = os.path.relpath(root, INPUT_DIR)
            out_xl = os.path.join(OUTPUT_DIR, rel,
                                  os.path.splitext(fname)[0] + ".xlsx")
            tasks.append((in_csv, out_xl))

    if not tasks:
        print("没有找到任何 CSV。")
        return

    with Pool(PROCESSES) as pool:
        pool.map(worker, tasks)

if __name__ == '__main__':
    main()
